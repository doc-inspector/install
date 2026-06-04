#!/usr/bin/env python3
"""
server.py — DocInspector Online Sandbox Server
================================================
Local Flask backend that processes uploaded PDF files using the same
CLI binaries bundled with the DocInspector desktop application.

Supported operations:
  - sanitize      : Remove all metadata with ExifTool
  - watermark     : Add a text watermark with Ghostscript
  - encrypt       : Password-protect PDF with qpdf
  - redact        : Find & redact text keywords (black boxes) via pdftk / qpdf
  - flatten       : Flatten PDF to image-only (remove text layer) via Ghostscript
  - rebuild       : Rebuild / repair corrupt PDF via qpdf

Usage:
  pip install flask
  python server.py
  Then open http://localhost:5000 in your browser.
"""

import os
import re
import json
import uuid
import shutil
import subprocess
import threading
import time
from pathlib import Path

try:
    from flask import Flask, request, jsonify, send_file, send_from_directory
    from flask_cors import CORS
except ImportError:
    print("=" * 60)
    print("ERROR: Flask not found. Please install dependencies:")
    print("  pip install flask flask-cors")
    print("=" * 60)
    raise SystemExit(1)

# ─────────────────────────────────────────────────────────────
# PATH CONFIGURATION — auto-detects Linux (Docker/VPS) vs Windows
# ─────────────────────────────────────────────────────────────
import platform as _platform
import shutil as _shutil

BASE_DIR  = Path(__file__).parent.resolve()
IS_LINUX  = _platform.system() != "Windows"

if IS_LINUX:
    # ── Linux / Docker / Hetzner VPS ──────────────────────────
    # Tools are installed system-wide via apt:
    #   apt install exiftool qpdf ghostscript mupdf-tools pdftk
    def _which(name: str) -> Path:
        p = _shutil.which(name)
        return Path(p) if p else Path(name)  # fallback, will fail gracefully

    TOOLS_DIR = Path("/usr/bin")  # informational only
    EXIFTOOL  = _which("exiftool")
    QPDF      = _which("qpdf")
    GS        = _which("gs")
    MUPDF     = _which("mutool")
    PDFTK     = _which("pdftk")
else:
    # ── Windows — bundled .exe files next to the desktop app ──
    APP_DIR = Path(r"E:\Website online and local app new project\APP 2.2.1 REAL TIME WORK\DocInspector_legacy\DocInspectorCSharp\bin\Debug\net8.0-windows\win-x64")
    TOOLS_DIR = APP_DIR / "Tools"
    EXIFTOOL  = TOOLS_DIR / "exiftool-13.43_64"  / "exiftool.exe"
    QPDF      = TOOLS_DIR / "qpdf-12.2.0-msvc64" / "bin" / "qpdf.exe"
    GS        = TOOLS_DIR / "gs"                 / "gswin64c.exe"
    MUPDF     = TOOLS_DIR / "mutool.exe"
    PDFTK     = None # Not bundled in newer versions

# Temp processing directory (auto-cleaned after each job)
TMP_DIR = BASE_DIR / "tmp_processing"
TMP_DIR.mkdir(exist_ok=True)

# Maximum file size per file for process endpoint
MAX_FILE_BYTES       = 50 * 1024 * 1024   # 50 MB for process
MAX_FILES_BATCH      = 100                 # 100 files per batch

# Bundle summary specific limits
BUNDLE_MAX_FILE_BYTES = 15 * 1024 * 1024  # 15 MB per file
BUNDLE_MAX_FILES      = 100               # 100 files per bundle

# ─────────────────────────────────────────────────────────────
# FLASK APP
# ─────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder=str(BASE_DIR), static_url_path="")
CORS(app, resources={r"/api/*": {"origins": "*"}})


# ─────────────────────────────────────────────────────────────
# STATIC FILE SERVING
# ─────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(BASE_DIR / "en", "index.html")

@app.route("/en/<path:filename>")
def serve_en(filename):
    return send_from_directory(BASE_DIR / "en", filename)

@app.route("/tools-online")
@app.route("/tools-online.html")
def tools_online():
    return send_from_directory(BASE_DIR / "en", "tools-online.html")

# ─────────────────────────────────────────────────────────────
# LICENSING & RATE LIMITING
# ─────────────────────────────────────────────────────────────
from functools import wraps
import urllib.request, urllib.parse, urllib.error
from collections import defaultdict
import datetime

# In-memory storage: resets on server restart
# USAGE_DB structure: { "identifier": { "date": "YYYY-MM-DD", "count": 0 } }
USAGE_DB = defaultdict(lambda: {"date": "", "count": 0})

def get_today():
    return datetime.date.today().isoformat()

def increment_usage(identifier: str, count: int = 1) -> int:
    today = get_today()
    record = USAGE_DB[identifier]
    if record["date"] != today:
        record["date"] = today
        record["count"] = 0
    record["count"] += count
    return record["count"]

def get_usage(identifier: str) -> int:
    today = get_today()
    record = USAGE_DB[identifier]
    if record["date"] != today:
        return 0
    return record["count"]

def validate_lemonsqueezy(license_key: str, activate=False) -> dict:
    """Validate or Activate with LemonSqueezy API. Returns tier ('Basic', 'Pro') or None."""
    endpoint = "activate" if activate else "validate"
    url = f"https://api.lemonsqueezy.com/v1/licenses/{endpoint}"
    data = urllib.parse.urlencode({
        "license_key": license_key,
        "instance_name": f"Web-{license_key[:8]}"
    }).encode("utf-8")
    
    req = urllib.request.Request(url, data=data, headers={"Accept": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            res_data = json.loads(response.read().decode())
            if res_data.get("valid") or res_data.get("activated") or "already active" in str(res_data.get("error","")).lower():
                meta = res_data.get("meta", {})
                variant = str(meta.get("variant_name", "")).lower()
                return {"tier": "Basic" if "basic" in variant else "Pro"}
    except Exception as e:
        print(f"LemonSqueezy {endpoint} error: {e}")
    return None

def get_client_ip():
    if request.headers.get("X-Forwarded-For"):
        return request.headers.get("X-Forwarded-For").split(",")[0].strip()
    return request.remote_addr

def get_current_tier():
    license_key = request.cookies.get("di_license")
    client_uuid = request.cookies.get("di_uuid")
    
    if license_key and client_uuid:
        record = USAGE_DB.get(license_key)
        # Check if this UUID is the currently active one for this license
        if record and record.get("active_web_uuid") == client_uuid:
            # Check 12-hour revalidation
            if time.time() - record.get("last_validated", 0) > 12 * 3600:
                print(f"Re-validating license {license_key[:8]}...")
                val = validate_lemonsqueezy(license_key, activate=False)
                if val:
                    record["last_validated"] = time.time()
                    record["tier"] = val["tier"]
                else:
                    # Revoked or expired! Kick them out.
                    del USAGE_DB[license_key]
                    return {"tier": "Free", "identifier": get_client_ip(), "revoked": True}
            
            return {"tier": record.get("tier", "Pro"), "identifier": license_key}
        else:
            # Another UUID took over this license (Kicked out!)
            return {"tier": "Free", "identifier": get_client_ip(), "kicked_out": True}
            
    return {"tier": "Free", "identifier": get_client_ip()}

@app.route("/api/auth/activate", methods=["POST"])
def auth_activate():
    key = request.form.get("license_key", "").strip()
    if not key:
        return jsonify({"error": "No license key provided."}), 400
    
    result = validate_lemonsqueezy(key, activate=True)
    if result:
        tier = result["tier"]
        
        # Generate new browser UUID for this session
        new_uuid = str(uuid.uuid4())
        
        # Take over the web slot for this license
        if key not in USAGE_DB:
            USAGE_DB[key] = {"count": 0, "date": get_today()}
        USAGE_DB[key]["active_web_uuid"] = new_uuid
        USAGE_DB[key]["last_validated"] = time.time()
        USAGE_DB[key]["tier"] = tier
        
        resp = jsonify({"status": "success", "tier": tier})
        resp.set_cookie("di_license", key, httponly=True, max_age=30*24*60*60, samesite="Lax")
        resp.set_cookie("di_tier", tier, max_age=30*24*60*60, samesite="Lax")
        resp.set_cookie("di_uuid", new_uuid, httponly=True, max_age=30*24*60*60, samesite="Lax")
        return resp
        
    return jsonify({"error": "Invalid or expired license key."}), 401

@app.route("/api/auth/status", methods=["GET"])
def auth_status():
    info = get_current_tier()
    tier = info["tier"]
    
    resp = jsonify({
        "tier": tier,
        "used_today": get_usage(info["identifier"]),
        "limit_today": "unlimited" if tier == "Pro" else (500 if tier == "Basic" else 100),
        "kicked_out": info.get("kicked_out", False),
        "revoked": info.get("revoked", False)
    })
    
    # If the user was kicked out by another browser, or license revoked, clear their cookies
    if info.get("kicked_out") or info.get("revoked"):
        resp.set_cookie("di_license", "", expires=0)
        resp.set_cookie("di_tier", "", expires=0)
        resp.set_cookie("di_uuid", "", expires=0)
        
    return resp

@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    resp = jsonify({"status": "logged_out"})
    resp.delete_cookie("di_license")
    resp.delete_cookie("di_tier")
    return resp


# ─────────────────────────────────────────────────────────────
# HEALTH CHECK
# ─────────────────────────────────────────────────────────────
@app.route("/api/health")
def health():
    tools = {
        "exiftool": EXIFTOOL.exists(),
        "qpdf":     QPDF.exists(),
        "ghostscript": GS.exists(),
        "mupdf":    MUPDF.exists(),
        "pdftk":    PDFTK.exists(),
    }
    return jsonify({"status": "ok", "tools": tools})


# ─────────────────────────────────────────────────────────────
# MAIN PROCESS ENDPOINT
# ─────────────────────────────────────────────────────────────
@app.route("/api/process", methods=["POST"])
def process():
    """
    Accepts multipart/form-data:
      files[]     : one or more PDF files (max 10)
      operations  : JSON array of operation names, e.g. ["sanitize","watermark"]
      options     : JSON object with per-operation settings
    """
    # --- License & Tier Check ---
    tier_info = get_current_tier()
    tier = tier_info["tier"]
    identifier = tier_info["identifier"]
    
    # --- Validate files ---
    uploaded = request.files.getlist("files[]")
    if not uploaded or len(uploaded) == 0:
        return jsonify({"error": "No files uploaded."}), 400
    
    num_files = len(uploaded)
    if num_files > MAX_FILES_BATCH:
        return jsonify({"error": f"Maximum {MAX_FILES_BATCH} files per batch."}), 400
        
    # Check rate limit
    current_usage = get_usage(identifier)
    if tier == "Free" and (current_usage + num_files > 100):
        return jsonify({"error": f"Free tier limit exceeded (100 files/day). You have used {current_usage}."}), 429
    elif tier == "Basic" and (current_usage + num_files > 500):
        return jsonify({"error": f"Basic tier limit exceeded (500 files/day). You have used {current_usage}."}), 429

    # --- Parse operations ---
    try:
        operations = json.loads(request.form.get("operations", "[]"))
    except Exception:
        return jsonify({"error": "Invalid operations JSON."}), 400

    if not operations:
        return jsonify({"error": "No operations selected."}), 400
        
    # Remove duplicate flatten if redact is present (since redact does flattening anyway)
    if "redact" in operations and "flatten" in operations:
        operations.remove("flatten")
        
    # Enforce logical execution order
    ORDER = {
        "rebuild": 1,
        "redact": 2,
        "flatten": 3,
        "watermark": 4,
        "sanitize": 5,
        "encrypt": 6
    }
    operations = sorted(operations, key=lambda x: ORDER.get(x, 99))
        
    # Feature gating
    pro_features = set()
    if tier == "Free":
        for op in operations:
            if op in pro_features:
                return jsonify({"error": f"Operation '{op}' requires a Basic or Pro license."}), 403

    # --- Parse options ---
    try:
        options = json.loads(request.form.get("options", "{}"))
    except Exception:
        options = {}

    # --- Create a unique job workspace ---
    job_id    = uuid.uuid4().hex
    job_dir   = TMP_DIR / job_id
    in_dir    = job_dir / "input"
    out_dir   = job_dir / "output"
    in_dir.mkdir(parents=True)
    out_dir.mkdir(parents=True)

    results   = []
    zip_path  = None

    try:
        for f in uploaded:
            if not f.filename:
                continue
            fname = Path(f.filename).name  # sanitize path
            in_path  = in_dir  / fname
            out_path = out_dir / f"processed_{fname}"

            # Validate size based on tier
            f.seek(0, 2)
            size = f.tell()
            f.seek(0)
            
            max_size_allowed = MAX_FILE_BYTES if tier in ["Basic", "Pro"] else (15 * 1024 * 1024)
            if size > max_size_allowed:
                results.append({"file": fname, "status": "error",
                                 "message": f"File exceeds {max_size_allowed//1024//1024} MB limit for {tier} tier."})
                continue

            # Save input
            f.save(str(in_path))

            # Run pipeline
            try:
                current_path = in_path
                for op in operations:
                    next_path = out_dir / f"_tmp_{op}_{fname}"
                    run_operation(op, current_path, next_path, options)
                    current_path = next_path

                # Final rename
                shutil.copy2(str(current_path), str(out_path))
                results.append({"file": fname, "status": "ok",
                                 "output": f"processed_{fname}"})
            except Exception as e:
                results.append({"file": fname, "status": "error",
                                 "message": str(e)})

        # --- Package outputs into a zip ---
        import zipfile
        zip_path = job_dir / f"docinspector_results_{job_id[:8]}.zip"
        
        success_count = 0
        with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
            error_log = []
            for r in results:
                if r["status"] == "ok":
                    fp = out_dir / r["output"]
                    if fp.exists():
                        zf.write(str(fp), arcname=r["output"])
                        success_count += 1
                elif r["status"] == "error":
                    fp_err = in_dir / r["file"]
                    if fp_err.exists():
                        zf.write(str(fp_err), arcname=f"errors/{r['file']}")
                    error_log.append(f"File: {r['file']} - Error: {r['message']}")
            
            if error_log:
                err_log_path = job_dir / "error_log.txt"
                err_log_path.write_text("\n".join(error_log), encoding="utf-8")
                zf.write(str(err_log_path), arcname="errors/error_log.txt")
                        
        if success_count > 0:
            increment_usage(identifier, success_count)

        # Schedule cleanup in 5 minutes
        threading.Thread(target=_delayed_cleanup,
                         args=(job_dir, 300), daemon=True).start()

        return send_file(
            str(zip_path),
            as_attachment=True,
            download_name=f"docinspector_results_{job_id[:8]}.zip",
            mimetype="application/zip"
        )

    except Exception as e:
        _cleanup(job_dir)
        return jsonify({"error": f"Server error: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# PDF BUNDLE SUMMARY ENDPOINT
# ─────────────────────────────────────────────────────────────
@app.route("/api/bundle-summary", methods=["POST"])
def bundle_summary():
    """
    Accepts multipart/form-data:
      files[]  : up to 100 PDF files (max 15 MB each)
    Returns:
      ZIP containing:
        - bundle_summary_<date>.xlsx   (Excel report)
        - bundle_summary_<date>.pdf    (PDF version)
    """
    # --- License & Tier Check ---
    tier_info = get_current_tier()
    tier = tier_info["tier"]
    identifier = tier_info["identifier"]

    uploaded = request.files.getlist("files[]")
    if not uploaded or len(uploaded) == 0:
        return jsonify({"error": "No files uploaded."}), 400
        
    num_files = len(uploaded)
    if num_files > BUNDLE_MAX_FILES:
        return jsonify({"error": f"Maximum {BUNDLE_MAX_FILES} files per bundle."}), 400
        
    # Check rate limit
    current_usage = get_usage(identifier)
    if tier == "Free" and (current_usage + num_files > 100):
        return jsonify({"error": f"Free tier limit exceeded (100 files/day). You have used {current_usage}."}), 429
    elif tier == "Basic" and (current_usage + num_files > 500):
        return jsonify({"error": f"Basic tier limit exceeded (500 files/day). You have used {current_usage}."}), 429

    job_id  = uuid.uuid4().hex
    job_dir = TMP_DIR / job_id
    in_dir  = job_dir / "input"
    out_dir = job_dir / "output"
    in_dir.mkdir(parents=True)
    out_dir.mkdir(parents=True)

    try:
        import hashlib
        import datetime

        # Try to import openpyxl for Excel generation
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            HAS_OPENPYXL = True
        except ImportError:
            HAS_OPENPYXL = False

        rows = []
        skipped = []

        for idx, f in enumerate(uploaded, start=1):
            if not f.filename:
                continue
            fname = Path(f.filename).name

            # Size check
            f.seek(0, 2)
            size = f.tell()
            f.seek(0)

            if size > BUNDLE_MAX_FILE_BYTES:
                skipped.append({"file": fname, "reason": f"Exceeds 15 MB ({size//1024//1024} MB)"})
                continue

            in_path = in_dir / fname
            f.save(str(in_path))

            # Extract metadata
            meta = _extract_pdf_meta(in_path)
            sha  = _sha256(in_path)

            rows.append({
                "nr":       idx,
                "name":     fname,
                "size_kb":  round(size / 1024, 1),
                "pages":    meta.get("pages", "—"),
                "author":   meta.get("author", "—"),
                "creator":  meta.get("creator", "—"),
                "created":  meta.get("created", "—"),
                "modified": meta.get("modified", "—"),
                "encrypted":meta.get("encrypted", False),
                "sha256":   sha,
            })

        if not rows:
            _cleanup(job_dir)
            return jsonify({"error": "No valid PDF files were processed."}), 400

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = out_dir / f"bundle_summary_{ts}.xlsx"
        pdf_path  = out_dir / f"bundle_summary_{ts}.pdf"

        if HAS_OPENPYXL:
            _generate_excel_report(rows, skipped, xlsx_path, is_pro=(tier=="Pro"))
        else:
            # Fallback: generate a plain CSV if openpyxl not available
            csv_path = out_dir / f"bundle_summary_{ts}.csv"
            _generate_csv_report(rows, csv_path, is_pro=(tier=="Pro"))
            xlsx_path = csv_path
            
        # Try to convert Excel → PDF using LibreOffice or skip
        _try_convert_to_pdf(xlsx_path, pdf_path)
        
        # Increment usage for successful files
        increment_usage(identifier, len(rows))

        # Package into ZIP
        import zipfile
        zip_path = job_dir / f"bundle_summary_{ts}.zip"
        with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
            if xlsx_path.exists():
                zf.write(str(xlsx_path), arcname=xlsx_path.name)
            if pdf_path.exists():
                zf.write(str(pdf_path), arcname=pdf_path.name)
            # Write a quick summary TXT always (fallback)
            txt_path = out_dir / f"bundle_summary_{ts}.txt"
            _generate_txt_report(rows, skipped, txt_path)
            zf.write(str(txt_path), arcname=txt_path.name)

        threading.Thread(target=_delayed_cleanup, args=(job_dir, 300), daemon=True).start()

        return send_file(
            str(zip_path),
            as_attachment=True,
            download_name=f"bundle_summary_{ts}.zip",
            mimetype="application/zip"
        )

    except Exception as e:
        _cleanup(job_dir)
        return jsonify({"error": f"Bundle summary error: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# BUNDLE SUMMARY HELPERS
# ─────────────────────────────────────────────────────────────

def _sha256(path: Path) -> str:
    """Compute SHA-256 of a file — mirrors C# SHA256.ComputeHash()."""
    import hashlib
    h = hashlib.sha256()
    try:
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(65536), b""):
                h.update(chunk)
        return h.hexdigest().upper()
    except Exception:
        return ""


def _is_pdf_protected(path: Path) -> bool:
    """
    Replicates C# BorderouEngine.IsFileProtected():
    - Detects PKCS#7 container (0x30 0x82 header)
    - Runs pdfinfo and checks 'Encrypted: yes'
    - Scans PDF byte stream for crypto markers:
      /Encrypt, /Type /Sig, /SigFlags, /DocMDP, /UR3,
      /Perms, /ByteRange, /Reference [, /TransformMethod
    """
    import re as _re
    try:
        with open(path, "rb") as fh:
            header = fh.read(8)
        # PKCS#7 signed container
        if len(header) >= 2 and header[0] == 0x30 and header[1] == 0x82:
            return True
        is_pdf = header[:5] == b"%PDF-"
        if not is_pdf:
            return False

        # pdfinfo check
        PDFINFO = _find_pdfinfo()
        if PDFINFO:
            try:
                r = subprocess.run([str(PDFINFO), str(path)],
                                   capture_output=True, text=True, timeout=10)
                out = r.stdout + r.stderr
                if _re.search(r"Encrypted:\s+yes", out, _re.IGNORECASE): return True
                if _re.search(r"changing:\s+not allowed", out, _re.IGNORECASE): return True
                if "password" in out.lower() or "encrypted" in r.stderr.lower(): return True
            except Exception:
                pass

        # byte-level scan (same as C# full file or last+first 512KB)
        size = path.stat().st_size
        patterns = [
            rb"/Type\s*/Sig", rb"/SigFlags", rb"/DocMDP", rb"/UR3",
            rb"/Perms", rb"/Encrypt", rb"/ByteRange", rb"/TransformMethod",
        ]
        with open(path, "rb") as fh:
            if size < 2 * 1024 * 1024:
                content = fh.read()
            else:
                content = fh.read(524288)  # first 512 KB
                fh.seek(-524288, 2)
                content += fh.read(524288)  # last 512 KB
        for pat in patterns:
            if _re.search(pat, content, _re.IGNORECASE):
                return True
        if b"Digitally signed" in content or b"Signature" in content:
            return True
    except Exception:
        pass
    return False


def _find_pdfinfo() -> Path | None:
    """Locate pdfinfo.exe in the tools directory."""
    candidates = [
        TOOLS_DIR / "poppler" / "bin" / "pdfinfo.exe",
        QPDF.parent / "pdfinfo.exe",
        TOOLS_DIR / "pdfinfo.exe",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _get_pdf_pages_fast(path: Path) -> int | None:
    """
    Get page count via pdfinfo first, then regex fallback —
    mirrors C# BorderouEngine.GetPdfPages().
    """
    import re as _re
    pdfinfo = _find_pdfinfo()
    if pdfinfo:
        try:
            r = subprocess.run([str(pdfinfo), str(path)],
                               capture_output=True, text=True, timeout=10)
            m = _re.search(r"^Pages:\s+(\d+)", r.stdout, _re.MULTILINE)
            if m:
                return int(m.group(1))
        except Exception:
            pass
    # ExifTool fallback
    if EXIFTOOL.exists():
        try:
            r = subprocess.run(
                [str(EXIFTOOL), "-j", "-PageCount", "-Pages", str(path)],
                capture_output=True, text=True, timeout=10)
            d = json.loads(r.stdout)
            if d:
                pg = d[0].get("PageCount") or d[0].get("Pages")
                if pg:
                    return int(pg)
        except Exception:
            pass
    # Regex byte scan fallback (mirrors C# regex fallback)
    try:
        with open(path, "rb") as fh:
            content = fh.read()
        count  = len(_re.findall(rb"/Type\s*/Page\b", content))
        count -= len(_re.findall(rb"/Type\s*/Pages\b", content))
        return count if count > 0 else 1
    except Exception:
        return None


def _extract_pdf_meta(path: Path) -> dict:
    """Extract PDF metadata using pdfinfo (qpdf fallback)."""
    meta = {}
    # Try pdfinfo (bundled with poppler tools or standalone)
    PDFINFO = TOOLS_DIR / "poppler" / "bin" / "pdfinfo.exe"
    if not PDFINFO.exists():
        # Try next to qpdf
        PDFINFO = QPDF.parent / "pdfinfo.exe"
    if not PDFINFO.exists():
        PDFINFO = None

    if PDFINFO and PDFINFO.exists():
        try:
            result = subprocess.run(
                [str(PDFINFO), str(path)],
                capture_output=True, text=True, timeout=15
            )
            for line in result.stdout.splitlines():
                lower = line.lower()
                if lower.startswith("pages:"):
                    try: meta["pages"] = int(line.split(":", 1)[1].strip())
                    except: pass
                elif lower.startswith("author:"):
                    meta["author"] = line.split(":", 1)[1].strip() or "—"
                elif lower.startswith("creator:"):
                    meta["creator"] = line.split(":", 1)[1].strip() or "—"
                elif lower.startswith("creationdate:"):
                    meta["created"] = line.split(":", 1)[1].strip()
                elif lower.startswith("moddate:"):
                    meta["modified"] = line.split(":", 1)[1].strip()
                elif "encrypted:" in lower:
                    meta["encrypted"] = "yes" in lower
        except Exception:
            pass

    # Try ExifTool as fallback
    if EXIFTOOL.exists() and ("pages" not in meta or "author" not in meta):
        try:
            result = subprocess.run(
                [str(EXIFTOOL), "-j", "-charset", "utf8",
                 "-Author", "-Creator", "-Pages", "-PageCount",
                 "-CreateDate", "-ModifyDate", "-Encryption",
                 str(path)],
                capture_output=True, text=True, timeout=15
            )
            data = json.loads(result.stdout)
            if data:
                d = data[0]
                if "pages" not in meta:
                    meta["pages"] = d.get("PageCount") or d.get("Pages") or "—"
                if "author" not in meta or meta.get("author") == "—":
                    meta["author"] = d.get("Author") or d.get("Creator") or "—"
                if "creator" not in meta or meta.get("creator") == "—":
                    meta["creator"] = d.get("Creator") or d.get("Producer") or "—"
                if "created" not in meta:
                    meta["created"] = d.get("CreateDate", "—")
                if "modified" not in meta:
                    meta["modified"] = d.get("ModifyDate", "—")
                if "encrypted" not in meta:
                    enc = d.get("Encryption", "")
                    meta["encrypted"] = bool(enc and enc.lower() not in ("", "none", "not encrypted"))
        except Exception:
            pass

    return meta


def _generate_excel_report(rows: list, skipped: list, out_path: Path, is_pro: bool = False):
    """
    Generates an Excel workbook with TWO sheets, exactly matching the C# desktop app:

    Sheet 1 — Basic Summary (BorderouEngine.cs):
      Columns: NR | File Name | Nr. of pages
      TableStyleMedium9, page break every 45 rows, portrait A4
      Footer: Total Files / Total Pages  (blue #376092 labels)

    Sheet 2 — Advanced Report (AdvancedBorderouEngine.cs):
      Columns: Nr | File Name | Format | Size Bytes | SHA256 | Protected | Nr. of pages | Last Write Time
      TableStyleMedium9, page break every 45 rows, landscape A3
      Footer: Total Files / Total Pages  (blue #376092 labels)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    HDR_BLUE  = "376092"   # matches C# XLColor.FromHtml("#376092")
    WHITE     = "FFFFFF"
    thin_side = Side(style="thin", color="BBBBBB")
    border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_pages = sum(r["pages"] if isinstance(r["pages"], int) else 0 for r in rows)

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════
    # SHEET 1 — Basic Summary (mirrors BorderouEngine.cs)
    # Columns: NR | File Name | Nr. of pages
    # Portrait A4, TableStyleMedium9
    # ═══════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"

    # Headers (row 1)
    for ci, (h, w) in enumerate(zip(["NR", "File Name", "Nr. of pages"], [8, 85, 15]), start=1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        c.fill = PatternFill("solid", fgColor=HDR_BLUE)
        c.alignment = center_al
        c.border = border
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    row_idx = 2
    for row in rows:
        ws1.cell(row_idx, 1, row["nr"]).alignment        = center_al
        ws1.cell(row_idx, 2, row["name"]).alignment      = left_al
        ws1.cell(row_idx, 2).style.alignment.wrap_text   = True
        pg = row["pages"]
        if isinstance(pg, int):
            ws1.cell(row_idx, 3, pg).alignment = center_al
        # Page break every 45 rows (mirrors C#)
        if row_idx > 1 and row_idx % 45 == 0:
            ws1.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=row_idx))
        row_idx += 1

    # Table (TableStyleMedium9)
    if len(rows) > 0:
        tbl_ref = f"A1:C{row_idx - 1}"
        from openpyxl.worksheet.table import Table, TableStyleInfo
        tbl1 = Table(displayName="SummaryTable", ref=tbl_ref)
        tbl1.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws1.add_table(tbl1)

    # Footer: Total Files / Total Pages (blue label, value in col 2)
    footer = row_idx + 1
    for offset, (lbl, val) in enumerate([("Total Files:", len(rows)), ("Total Pages:", total_pages)]):
        lc = ws1.cell(footer + offset, 1, lbl)
        lc.fill      = PatternFill("solid", fgColor=HDR_BLUE)
        lc.font      = Font(bold=True, color=WHITE)
        lc.alignment = Alignment(horizontal="right")
        lc.border    = border
        vc = ws1.cell(footer + offset, 2, val)
        vc.font      = Font(bold=True)
        vc.alignment = left_al
        vc.fill      = PatternFill("solid", fgColor="FFFFFF")

    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.page_setup.orientation = ws1.ORIENTATION_PORTRAIT
    ws1.page_setup.paperSize   = ws1.PAPERSIZE_A4
    ws1.page_setup.fitToWidth  = 1
    ws1.print_title_rows       = "1:1"
    ws1.sheet_view.showGridLines = True

    # ═══════════════════════════════════════════════════
    # SHEET 2 — Advanced Report (Pro Tier only)
    # ═══════════════════════════════════════════════════
    if not is_pro:
        wb.save(str(out_path))
        return

    ws2 = wb.create_sheet("Advanced Report")

    adv_headers = ["Nr", "File Name", "Format", "Size Bytes", "SHA256",
                   "Protected", "Nr. of pages", "Last Write Time"]
    adv_widths  = [6, 80, 10, 15, 75, 12, 12, 20]
    for ci, (h, w) in enumerate(zip(adv_headers, adv_widths), start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_BLUE)
        c.alignment = center_al
        c.border = border
        ws2.column_dimensions[get_column_letter(ci)].width = w

    import os as _os
    row_idx2   = 2
    total_pg2  = 0
    for row in rows:
        fpath = row.get("_path")  # set during upload loop
        lw = ""
        if fpath and Path(fpath).exists():
            lw = datetime.datetime.fromtimestamp(
                Path(fpath).stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")

        pg = row["pages"]
        if isinstance(pg, int):
            total_pg2 += pg

        ws2.cell(row_idx2, 1, row["nr"]).alignment      = center_al
        ws2.cell(row_idx2, 2, row["name"]).alignment    = left_al
        ws2.cell(row_idx2, 2).style.alignment.wrap_text = True
        ext = Path(row["name"]).suffix.lstrip(".").lower()
        ws2.cell(row_idx2, 3, ext).alignment             = center_al
        ws2.cell(row_idx2, 4, int(row["size_kb"] * 1024)).alignment = center_al
        ws2.cell(row_idx2, 5, row["sha256"]).alignment  = left_al
        prot = "YES" if row["encrypted"] else "NO"
        pc = ws2.cell(row_idx2, 6, prot)
        pc.alignment = center_al
        if row["encrypted"]:
            pc.font = Font(name="Calibri", color="C0392B", bold=True, size=9)
        ws2.cell(row_idx2, 7, pg if isinstance(pg, int) else "").alignment = center_al
        ws2.cell(row_idx2, 8, lw).alignment = center_al

        if row_idx2 > 1 and row_idx2 % 45 == 0:
            ws2.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=row_idx2))
        row_idx2 += 1

    if len(rows) > 0:
        tbl_ref2 = f"A1:H{row_idx2 - 1}"
        tbl2 = Table(displayName="AdvancedTable", ref=tbl_ref2)
        tbl2.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws2.add_table(tbl2)

    footer2 = row_idx2 + 1
    for offset, (lbl, val) in enumerate([("Total Files:", len(rows)), ("Total Pages:", total_pg2)]):
        lc = ws2.cell(footer2 + offset, 1, lbl)
        lc.fill = PatternFill("solid", fgColor=HDR_BLUE)
        lc.font = Font(bold=True, color=WHITE)
        lc.alignment = Alignment(horizontal="right")
        lc.border = border
        vc = ws2.cell(footer2 + offset, 2, val)
        vc.font = Font(bold=True)
        vc.alignment = left_al

    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize   = ws2.PAPERSIZE_A3
    ws2.page_setup.fitToWidth  = 1
    ws2.print_title_rows       = "1:1"

    # Skipped files note at bottom of sheet 2
    if skipped:
        sk = footer2 + 3
        ws2.cell(sk, 1, "SKIPPED (too large):")\
           .font = Font(bold=True, color="E67E22")
        for i, s in enumerate(skipped):
            ws2.cell(sk + 1 + i, 1, s["file"])
            ws2.cell(sk + 1 + i, 2, s["reason"])

    wb.save(str(out_path))


def _generate_csv_report(rows: list, out_path: Path, is_pro: bool = False):
    """CSV fallback if openpyxl not available."""
    import datetime
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("Nr,Name,Size_KB,Pages,Author,Creator,Created,Modified,Encrypted,SHA256\n")
        for r in rows:
            enc = "YES" if r["encrypted"] else "No"
            f.write(f'{r["nr"]},"{r["name"]}",{r["size_kb"]},{r["pages"]},"{r["author"]}","{r["creator"]}","{r["created"]}","{r["modified"]}",{enc},"{r["sha256"]}"\n')


def _generate_txt_report(rows: list, skipped: list, out_path: Path):
    """Plain text summary always included in ZIP as a quick reference."""
    import datetime
    total_pages  = sum(r["pages"] if isinstance(r["pages"], int) else 0 for r in rows)
    total_enc    = sum(1 for r in rows if r["encrypted"])
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("  DocInspector — PDF Bundle Summary\n")
        f.write(f"  Generated : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"  Total files: {len(rows)}   Total pages: {total_pages}   Encrypted: {total_enc}\n")
        f.write("=" * 70 + "\n\n")
        for r in rows:
            enc = "[ENCRYPTED]" if r["encrypted"] else ""
            f.write(f"  {r['nr']:>3}. {r['name']:<45} {r['size_kb']:>8} KB   {str(r['pages']):>4} pgs  {enc}\n")
            f.write(f"       Author   : {r['author']}\n")
            f.write(f"       Creator  : {r['creator']}\n")
            f.write(f"       SHA-256  : {r['sha256'][:32]}...\n\n")
        if skipped:
            f.write("\nSKIPPED (too large):\n")
            for s in skipped:
                f.write(f"  - {s['file']} : {s['reason']}\n")


def _try_convert_to_pdf(xlsx_path: Path, pdf_path: Path):
    """Try to convert Excel to PDF using LibreOffice (if available)."""
    try:
        import shutil as _sh
        lo = _sh.which("soffice") or _sh.which("libreoffice")
        if lo:
            subprocess.run(
                [lo, "--headless", "--convert-to", "pdf",
                 "--outdir", str(xlsx_path.parent), str(xlsx_path)],
                capture_output=True, timeout=60
            )
            # LibreOffice outputs <name>.pdf next to the xlsx
            candidate = xlsx_path.with_suffix(".pdf")
            if candidate.exists() and not pdf_path.exists():
                candidate.rename(pdf_path)
    except Exception:
        pass  # PDF conversion is optional — Excel is always present




# ─────────────────────────────────────────────────────────────
# OPERATION DISPATCHER
# ─────────────────────────────────────────────────────────────
def run_operation(op: str, in_path: Path, out_path: Path, options: dict):
    """Route to the correct CLI handler."""
    op = op.lower().strip()
    if op == "sanitize":
        op_sanitize(in_path, out_path, options)
    elif op == "watermark":
        op_watermark(in_path, out_path, options)
    elif op == "encrypt":
        op_encrypt(in_path, out_path, options)
    elif op == "redact":
        op_redact(in_path, out_path, options)
    elif op == "flatten":
        op_flatten(in_path, out_path)
    elif op == "rebuild":
        op_rebuild(in_path, out_path)
    else:
        raise ValueError(f"Unknown operation: {op}")


# ─────────────────────────────────────────────────────────────
# OPERATION IMPLEMENTATIONS
# ─────────────────────────────────────────────────────────────

def op_sanitize(in_path: Path, out_path: Path, options: dict = None):
    """Strip metadata from PDF using ExifTool based on options."""
    _require_tool(EXIFTOOL, "ExifTool")
    shutil.copy2(str(in_path), str(out_path))
    
    cmd = [str(EXIFTOOL), "-overwrite_original"]
    
    if options:
        sanitize_author = options.get("sanitizeAuthor", True)
        sanitize_time = options.get("sanitizeTime", True)
        sanitize_gps = options.get("sanitizeGps", True)
        sanitize_software = options.get("sanitizeSoftware", True)
        
        # If all options are True, do a full deep clean (default)
        if sanitize_author and sanitize_time and sanitize_gps and sanitize_software:
            cmd.append("-all=")
        else:
            if sanitize_author:
                cmd.extend(["-Author=", "-Title=", "-Creator=", "-Subject="])
            if sanitize_time:
                cmd.extend(["-CreateDate=", "-ModifyDate=", "-MetadataDate=", "-History="])
            if sanitize_gps:
                cmd.extend(["-GPSLatitude=", "-GPSLongitude=", "-GPSPosition=", "-GPSLatitudeRef=", "-GPSLongitudeRef="])
            if sanitize_software:
                cmd.extend(["-Software=", "-CreatorTool=", "-Producer="])
    else:
        # Full deep clean by default
        cmd.append("-all=")
        
    cmd.append(str(out_path))
    _run(cmd)


def op_watermark(in_path: Path, out_path: Path, options: dict):
    """Add a watermark to all pages using pypdf and reportlab for true transparency."""
    import io
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import Color

    text     = options.get("watermarkText",    "CONFIDENTIAL")
    font_sz  = int(options.get("watermarkSize",    60))
    opacity  = float(options.get("watermarkOpacity", 0.15))
    color_str= options.get("watermarkColor",   "0.5 0.5 0.5")  # R G B 0-1
    angle    = float(options.get("watermarkAngle", -45))

    # Parse color
    r, g, b = [float(c) for c in color_str.split()]
    draw_color = Color(r, g, b, alpha=opacity)

    FONT_MAP = {
        "Helvetica-Bold":      "Helvetica-Bold",
        "Helvetica":           "Helvetica",
        "Times-Bold":          "Times-Bold",
        "Times-Roman":         "Times-Roman",
        "Courier-Bold":        "Courier-Bold",
        "Courier":             "Courier",
        "Helvetica-Oblique":   "Helvetica-Oblique",
        "Times-BoldItalic":    "Times-BoldItalic",
        "Palatino-Roman":      "Times-Roman", # fallback if not standard, but let's try mapping to Helvetica for now if unsupported
    }
    raw_font = options.get("watermarkFont", "Helvetica-Bold")
    gs_font  = FONT_MAP.get(raw_font, "Helvetica-Bold")
    if raw_font == "Palatino-Roman":
        # Reportlab supports standard fonts. Times-Roman is a close serif.
        gs_font = "Times-Roman"

    lines = [line.strip() for line in text.replace('\r', '').split('\n')]
    
    # Process PDF
    reader = PdfReader(str(in_path))
    writer = PdfWriter()

    for page in reader.pages:
        # Get page dimensions
        mbox = page.mediabox
        page_w = float(mbox.width)
        page_h = float(mbox.height)

        # Create transparent stamp in memory for this specific page size
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=(page_w, page_h))
        c.setFillColor(draw_color)
        c.setFont(gs_font, font_sz)

        # Move to center
        c.translate(page_w / 2.0, page_h / 2.0)
        c.rotate(angle)

        # Draw lines centered
        N = len(lines)
        for i, line in enumerate(lines):
            y_offset = ((N - 1) / 2.0 - i) * font_sz * 1.2
            c.drawCentredString(0, y_offset, line)

        c.save()
        packet.seek(0)
        
        # Merge
        stamp_reader = PdfReader(packet)
        stamp_page = stamp_reader.pages[0]
        page.merge_page(stamp_page, over=True)
        writer.add_page(page)

    with open(str(out_path), "wb") as fOut:
        writer.write(fOut)


def op_encrypt(in_path: Path, out_path: Path, options: dict):
    """Password-protect PDF using qpdf."""
    _require_tool(QPDF, "qpdf")
    user_pw  = options.get("encryptUserPw",  "")
    owner_pw = options.get("encryptOwnerPw", "")
    bits     = options.get("encryptBits",    "256")
    
    no_print = options.get("encryptNoPrint", False)
    no_copy  = options.get("encryptNoCopy", False)
    no_edit  = options.get("encryptNoEdit", False)

    # If any restriction is set, but no owner password is provided,
    # we must generate a random owner password so restrictions are actually enforced.
    import secrets
    if (no_print or no_copy or no_edit) and not owner_pw:
        owner_pw = secrets.token_hex(8)
    
    # If owner password is provided but no user password, qpdf requires user password to be empty string
    if not owner_pw:
        owner_pw = user_pw if user_pw else secrets.token_hex(8)

    cmd = [
        str(QPDF), "--encrypt",
        user_pw, owner_pw, bits
    ]
    
    if no_print:
        cmd.append("--print=none")
    if no_copy:
        cmd.append("--extract=n")
    if no_edit:
        cmd.append("--modify=none")
        
    cmd.extend(["--", str(in_path), str(out_path)])
    
    _run(cmd)


def op_watermark(in_path: Path, out_path: Path, options: dict):
    """Add a watermark to all pages using pypdf and reportlab for true transparency."""
    import io
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import Color

    text     = options.get("watermarkText",    "CONFIDENTIAL")
    font_sz  = int(options.get("watermarkSize",    60))
    opacity  = float(options.get("watermarkOpacity", 0.15))
    color_str= options.get("watermarkColor",   "0.5 0.5 0.5")  # R G B 0-1
    angle    = float(options.get("watermarkAngle", -45))

    # Parse color
    r, g, b = [float(c) for c in color_str.split()]
    draw_color = Color(r, g, b, alpha=opacity)

    FONT_MAP = {
        "Helvetica-Bold":      "Helvetica-Bold",
        "Helvetica":           "Helvetica",
        "Times-Bold":          "Times-Bold",
        "Times-Roman":         "Times-Roman",
        "Courier-Bold":        "Courier-Bold",
        "Courier":             "Courier",
        "Helvetica-Oblique":   "Helvetica-Oblique",
        "Times-BoldItalic":    "Times-BoldItalic",
        "Palatino-Roman":      "Times-Roman", # fallback if not standard, but let's try mapping to Helvetica for now if unsupported
    }
    raw_font = options.get("watermarkFont", "Helvetica-Bold")
    gs_font  = FONT_MAP.get(raw_font, "Helvetica-Bold")
    if raw_font == "Palatino-Roman":
        # Reportlab supports standard fonts. Times-Roman is a close serif.
        gs_font = "Times-Roman"

    lines = [line.strip() for line in text.replace('\r', '').split('\n')]
    
    # Process PDF
    reader = PdfReader(str(in_path))
    writer = PdfWriter()

    for page in reader.pages:
        # Get page dimensions
        mbox = page.mediabox
        page_w = float(mbox.width)
        page_h = float(mbox.height)

        # Create transparent stamp in memory for this specific page size
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=(page_w, page_h))
        c.setFillColor(draw_color)
        c.setFont(gs_font, font_sz)

        # Move to center
        c.translate(page_w / 2.0, page_h / 2.0)
        c.rotate(angle)

        # Draw lines centered
        N = len(lines)
        for i, line in enumerate(lines):
            y_offset = ((N - 1) / 2.0 - i) * font_sz * 1.2
            c.drawCentredString(0, y_offset, line)

        c.save()
        packet.seek(0)
        
        # Merge
        stamp_reader = PdfReader(packet)
        stamp_page = stamp_reader.pages[0]
        page.merge_page(stamp_page, over=True)
        writer.add_page(page)

    with open(str(out_path), "wb") as fOut:
        writer.write(fOut)


def op_encrypt(in_path: Path, out_path: Path, options: dict):
    """Password-protect PDF using qpdf."""
    _require_tool(QPDF, "qpdf")
    user_pw  = options.get("encryptUserPw",  "")
    owner_pw = options.get("encryptOwnerPw", options.get("encryptUserPw", "docprotect"))
    bits     = options.get("encryptBits",    "256")

    cmd = [
        str(QPDF), "--encrypt",
        user_pw, owner_pw, str(bits),
        "--", str(in_path), str(out_path)
    ]
    _run(cmd)


def op_redact(in_path: Path, out_path: Path, options: dict):
    """
    Find & redact keywords using PyMuPDF (fitz).
    It searches for the exact phrases and adds black rectangle redactions,
    removing the underlying text and vectors.
    """
    keywords_raw = options.get("redactKeywords", "")
    keywords = [k.strip() for k in keywords_raw.split(',') if k.strip()]
    
    if not keywords:
        # If no keywords are provided, fallback to flatten to image
        op_flatten(in_path, out_path)
        return

    try:
        import fitz
    except ImportError:
        # Fallback if PyMuPDF is not installed
        op_flatten(in_path, out_path)
        return

    doc = fitz.open(str(in_path))
    for page in doc:
        for kw in keywords:
            text_instances = page.search_for(kw)
            for inst in text_instances:
                page.add_redact_annot(inst, fill=(0, 0, 0))
        # physically apply redactions to remove underlying text/images
        page.apply_redactions()

    doc.save(str(out_path), garbage=4, deflate=True)
    doc.close()


def op_flatten(in_path: Path, out_path: Path):
    """
    Flatten PDF to image-only (no text layer) using Ghostscript.
    This renders every page as a raster image inside a new PDF.
    """
    _require_tool(GS, "Ghostscript")
    _run([
        str(GS), "-dBATCH", "-dNOPAUSE", "-dQUIET",
        "-sDEVICE=pdfimage24",
        "-r200",
        f"-sOutputFile={out_path}",
        str(in_path)
    ])


def op_rebuild(in_path: Path, out_path: Path):
    """
    Rebuild / linearize a PDF using qpdf to fix structural issues.
    """
    _require_tool(QPDF, "qpdf")
    _run([
        str(QPDF), "--linearize",
        str(in_path), str(out_path)
    ])


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def _require_tool(path: Path, name: str):
    if not path.exists():
        raise FileNotFoundError(
            f"{name} binary not found at: {path}\n"
            "Ensure the DocInspector desktop app is installed in the expected location."
        )


def _run(cmd: list):
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=120
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"CLI error (code {result.returncode}):\n"
            f"STDOUT: {result.stdout[:500]}\n"
            f"STDERR: {result.stderr[:500]}"
        )


def _delayed_cleanup(path: Path, delay_secs: int):
    time.sleep(delay_secs)
    _cleanup(path)


def _cleanup(path: Path):
    try:
        if path.exists():
            shutil.rmtree(str(path), ignore_errors=True)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────
# STARTUP
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  DocInspector Online Sandbox — Local Server")
    print("=" * 60)
    print(f"  Base dir  : {BASE_DIR}")
    print(f"  Tools dir : {TOOLS_DIR}")
    print()
    print("  Tool availability:")
    for name, path in [("ExifTool", EXIFTOOL), ("qpdf", QPDF),
                        ("Ghostscript", GS), ("MuPDF", MUPDF), ("pdftk", PDFTK)]:
        exists = path is not None and path.exists()
        status = "[OK] Found" if exists else "[FAIL] Missing"
        print(f"    {name:15s}: {status}")
    print()
    print("  Open in browser: http://localhost:5000/tools-online")
    print("  Press Ctrl+C to stop.")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
